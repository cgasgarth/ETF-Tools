import openpyxl

from FinClasses import ETF


def importFromExcel(fileName):
    workbook = openpyxl.load_workbook(filename = fileName, data_only = True)
    sheet = workbook.active
    etfList = []
    i = 4
    while (sheet.cell(row=i, column=1).value != None):
        ticker = sheet.cell(row=i, column=1).value
        weight = float(sheet.cell(row=i, column=8).value)
        if ticker == "CASH":
            break
        newETF = ETF(ticker, weight)
        etfList.append(newETF)
        i += 1
    return etfList
